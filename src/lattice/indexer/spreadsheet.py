"""Spreadsheet indexer using openpyxl.

Passage IDs: p.<sheet>.<cell_ref>, e.g. p.Sheet1.A1
Schema-like information (header row + summary stats) goes into the first
passage per sheet; individual non-empty cells become data_point passages.
Heavy numeric extraction is deferred — this is the minimal viable pass.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from ..graph.models import (
    Citation,
    Passage,
    PassageLocation,
    PassageType,
    Source,
    SourceMetadata,
    SourceType,
)
from .base import Indexer


class SpreadsheetIndexer(Indexer):
    def index(self, file_path: Path) -> Source:
        wb = load_workbook(filename=str(file_path), data_only=True, read_only=True)
        passages: list[Passage] = []

        for sheet in wb.worksheets:
            header = _read_header(sheet)
            if header:
                hdr_text = f"Columns: {', '.join(header)}"
                passages.append(
                    Passage(
                        id=f"p.{sheet.title}.header",
                        text=hdr_text,
                        location=PassageLocation(cell=f"{sheet.title}!1:1", section=sheet.title),
                        type=PassageType.claim,
                        char_count=len(hdr_text),
                    )
                )
            for row in sheet.iter_rows(min_row=2, values_only=False):
                for cell in row:
                    if cell.value is None:
                        continue
                    cell_ref = f"{sheet.title}!{cell.coordinate}"
                    text = str(cell.value)
                    passages.append(
                        Passage(
                            id=f"p.{sheet.title}.{cell.coordinate}",
                            text=text,
                            location=PassageLocation(cell=cell_ref, section=sheet.title),
                            type=PassageType.data_point,
                            char_count=len(text),
                        )
                    )

        return Source(
            source_id=Indexer.slugify(file_path.stem),
            type=SourceType.dataset,
            citation=Citation(authors=[], year=None, title=file_path.stem),
            passages=passages,
            metadata=SourceMetadata(
                date_added=datetime.now(timezone.utc),
                file_path=str(file_path),
                hash="",
            ),
        )


def _read_header(sheet) -> list[str]:
    try:
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        return []
    return [str(c) for c in first_row if c is not None]
